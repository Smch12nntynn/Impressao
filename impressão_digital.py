from tkinter import *
from tkcalendar import DateEntry
from tkinter import ttk
from tkinter import messagebox
import openpyxl
from openpyxl.styles import Alignment, Font
import os 
import random


root = Tk()

class Aplication():
    def __init__(self) -> None:
        self.book = openpyxl.Workbook()
        self.template_name = "Template.xlsx"
        self.BD_workbook = "Banco de Dados"
        self.monthly_workbook = "Mensal"
        self.analysis_workbook = "Analise"
        self.this_book_name = ""
        self.months = {
        1: "Janeiro",
        2: "Fevereiro",
        3: "Março",
        4: "Abril",
        5: "Maio",
        6: "Junho",
        7: "Julho",
        8: "Agosto",
        9: "Setembro",
        10: "Outubro",
        11: "Novembro",
        12: "Dezembro",
        }
    def check_table_existence(self, filename):
        return os.path.isfile(filename)
    def open_workbook(self, filename):
        workbook = openpyxl.load_workbook(filename)
        return workbook
    def open_BD_worksheet(self, workbook):
        aplication_class = Aplication()
        worksheet = workbook[aplication_class.BD_workbook]
        return worksheet
    def open_monthly_worksheet(self, workbook):
        aplication_class = Aplication()
        worksheet = workbook[aplication_class.monthly_workbook]
        return worksheet
    def fill_worksheet(self, worksheet, data, posicion) -> None:
        alignment = Alignment(horizontal="center", vertical="center")

        def for_cell(line):
            for i in range(1,10):
                cell = worksheet.cell(line,i,data[i-1])
                cell.alignment = alignment
                if i > 7:
                    cell.number_format = 'R$ 0.00'

        if posicion == 0:
            for_cell(1)
        else:
            for_cell(posicion + 1)           
    def name_worksheet(self):
        aplication_class = Aplication()
        try:
            month = str(aplication_class.months[int(self.entry_data.get().split("/")[0])])
            year = str(self.entry_data.get().split("/")[2])
            name = month + "_20" + year + ".xlsx"
            return name
        except ValueError:
            messagebox.showerror("Erro", "Esqueceu de preencher a DATA.")
    def data_structure(self):
        data = []
        data.append(self.entry_data.get())
        data.append(str(self.entry_hora.get()))
        data.append(int(self.entry_copias_br.get()))
        data.append(int(self.entry_copias_r.get()))
        data.append(int(self.entry_perdas_br.get()))
        data.append(int(self.entry_perdas_r.get()))
        data.append(float(self.entry_pg_dinheiro.get()))
        data.append(float(self.entry_pg_pix.get()))
        return data
    def get_id(self):
        id = str(self.entry_id.get())
        return id
    def clean_table(self):
        self.entry_id.config(state='normal')
        self.entry_id.delete(0, END)
        self.entry_data.delete(0, END)
        self.entry_hora.delete(0, END)
        self.entry_copias_br.delete(0, END)
        self.entry_copias_r.delete(0, END)
        self.entry_perdas_br.delete(0, END)
        self.entry_perdas_r.delete(0, END)
        self.entry_pg_dinheiro.delete(0, END)
        self.entry_pg_pix.delete(0, END)
        self.entry_id.delete(0, END)
    def insert_tree(self, sheet) -> None:
        rows = sheet.iter_rows()
        self.list_print.delete(*self.list_print.get_children())
        next(rows)
        rows = list(rows)[::-1]
        for row in rows:
            values = [cell.value for cell in row]
            self.list_print.insert('', 'end', values=values)
    def on_double_click(self, event):
        index = event.widget.selection()[0]
        item = event.widget.item(index)
        data = item['values']
        self.clean_table()
        self.insert_entry(data)
    def id_generator(self, size):
        caracteres = "abcdefghijklmnopqrstuvwxyz0123456789ABCDEFGHIJKLMNOPQRSTUVXYZ"
        return "".join(random.choice(caracteres) for _ in range(size))
    def get_row_data_from_id(self, sheet, id):
        column = sheet['A']
        try:
            for cell in column:
                row_value = cell.value
                if id in row_value:
                    row_cells = sheet[cell.row]
                    row_data = [cell.value for cell in row_cells]
                    return row_data
        except ValueError:
            messagebox.showerror("Erro", "O ID nao foi encontrado, certifique se esta digitando corretamente.")
    def get_row_index_from_id(self, id, sheet):
        for row_index, row in enumerate(sheet.iter_rows()):
            if row[0].value == id:
                return int(row_index)
        return -1
    def insert_entry(self, array):
        try:
            self.entry_id.insert(0, array[0])
            self.entry_id.config(state='readonly')
            self.entry_data.insert(0, array[1])
            self.entry_hora.insert(0, array[2])
            self.entry_copias_br.insert(0, array[3])
            self.entry_copias_r.insert(0, array[4])
            self.entry_perdas_br.insert(0, array[5])
            self.entry_perdas_r.insert(0, array[6])
            self.entry_pg_dinheiro.insert(0, array[7])
            self.entry_pg_pix.insert(0, array[8])
        except ValueError:
            messagebox.showerror("Erro", "O ID nao foi encontrado, certifique se esta digitando corretamente.")
    def find_excel_files(self, except_file_name="Template.xlsx"):
        path = os.getcwd()
        files = os.listdir(path)
        files = [file for file in files if file.endswith(".xlsx")]
        if except_file_name in files:
            files.remove(except_file_name)
        return files
    def get_selected_option(self, popup, listbox):
        selected_option = listbox.get(listbox.curselection()[0])
        popup.destroy()
        return selected_option
    def delete_row_by_id(self, name, id) -> None:
        wb = self.open_workbook(name)
        ws = self.open_BD_worksheet(wb)
        row_index = self.get_row_index_from_id(id, ws)
        if row_index >= 0:
            for cell in ws[row_index + 1]:
                cell.value = None 
        for row in ws.iter_rows():
            if not all(cell.value for cell in row):
                ws.delete_rows(row[0].row, 1)
        wb.save(name)
        self.clean_table()
    def make_monthly_data(self, bd_ws):
        coluna_data = 2
        coluna_copias_br = 4
        coluna_copias_ri = 5
        coluna_perdas_br = 6
        coluna_perdas_ri = 7
        coluna_dinheiro = 8
        coluna_pix = 9

        dados_dia = {}
        for row in range(2, bd_ws.max_row + 1):
            data_celula = bd_ws.cell(row=row, column=coluna_data).value
            copias_br = bd_ws.cell(row=row, column=coluna_copias_br).value
            copias_ri = bd_ws.cell(row=row, column=coluna_copias_ri).value
            perdas_br = bd_ws.cell(row=row, column=coluna_perdas_br).value
            perdas_ri = bd_ws.cell(row=row, column=coluna_perdas_ri).value
            dinheiro = bd_ws.cell(row=row, column=coluna_dinheiro).value
            pix = bd_ws.cell(row=row, column=coluna_pix).value
            data_objeto = data_celula.split("/")
            dia = data_objeto[1]
            if dia not in dados_dia:
                dados_dia[dia] = {
                    "copias_br": 0,
                    "copias_ri": 0,
                    "perdas_br": 0,
                    "perdas_ri": 0,
                    "dinheiro": 0,
                    "pix": 0,
                    }
            dados_dia[dia]["copias_br"] += copias_br
            dados_dia[dia]["copias_ri"] += copias_ri
            dados_dia[dia]["perdas_br"] += perdas_br
            dados_dia[dia]["perdas_ri"] += perdas_ri
            dados_dia[dia]["dinheiro"] += dinheiro
            dados_dia[dia]["pix"] += pix
        return dados_dia
    def extract_month_year(self, filename):
        name_without_ext = filename.split(".")[0]
        month, year = name_without_ext.split("_")
        month = month.capitalize()
        return f"{month} {year}"
    def insert_monthly_data(self, m_ws, data, name):
        worksheet = m_ws
        alignment = Alignment(horizontal="center", vertical="center")
        font = Font(bold=True, size=16)

        month_year = self.extract_month_year(name)
        worksheet.cell(1, 4, month_year).font = font
        worksheet.cell(1, 4, month_year).alignment = alignment
        for i in range(1, 32):
            line = i + 4
            day = str(worksheet.cell(line, 1).value)
            if day in data:                
                worksheet.cell(line, 2, data[day]['copias_br']).alignment = alignment                
                worksheet.cell(line, 3, data[day]['copias_ri']).alignment = alignment                
                worksheet.cell(line, 4, data[day]['perdas_br']).alignment = alignment                
                worksheet.cell(line, 5, data[day]['perdas_ri']).alignment = alignment                
                worksheet.cell(line, 6, data[day]['copias_br'] + data[day]['perdas_br']).alignment = alignment                
                worksheet.cell(line, 7, data[day]['copias_br'] + data[day]['perdas_ri']).alignment = alignment
                worksheet.cell(line, 8, data[day]['dinheiro']).alignment = alignment
                worksheet.cell(line, 9, data[day]['pix']).alignment = alignment

class Buttons(Aplication):
    def __init__(self) -> None:
        self.button_save()
        self.button_find()
        self.button_delete()
    def button_save(self) -> None:
        id = self.get_id()
        aplication_class = Aplication()
        data = self.data_structure()
        name = self.name_worksheet()

        if id == '':
            if not self.check_table_existence(name):
                wb = self.open_workbook(aplication_class.template_name)
                ws = self.open_BD_worksheet(wb)
                new_id = self.id_generator(4)
                data.insert(0, new_id)
                self.fill_worksheet(ws, data, 1)
                wb.save(name)
                self.this_book_name = name
                self.insert_tree(ws)
            else:
                wb = self.open_workbook(name)
                ws = self.open_BD_worksheet(wb)
                last_row = ws.max_row
                new_id = self.id_generator(4)
                data.insert(0, new_id)
                self.fill_worksheet(ws, data, last_row)
                wb.save(name)
                self.this_book_name = name
                self.insert_tree(ws)
        else:
            wb = self.open_workbook(name)
            ws = self.open_BD_worksheet(wb)
            column_values = [cell.value for cell in ws['A']]
            posicion = column_values.index(id)
            if posicion != -1:
                line = posicion
                data.insert(0, id)
                self.fill_worksheet(ws, data, line)
                wb.save(name)
                self.this_book_name = name
                self.insert_tree(ws)
            else:
                print("Algum erro ou id nao econtrado")
        self.clean_table()
    def button_find(self) -> None:
        id = self.get_id()
        name = self.this_book_name
        wb = self.open_workbook(name)
        ws = self.open_BD_worksheet(wb)
        try:
            row_data = self.get_row_data_from_id(ws, id)
            self.clean_table()
            self.insert_entry(row_data)
        except ValueError:
            return messagebox.showerror("Erro", "O ID nao foi encontrado, certifique se esta digitando corretamente.")
    def button_delete(self):
        id = self.entry_id.get()
        name = self.this_book_name
        if name != "":
            self.delete_row_by_id(name, id)
            self.insert_tree(self.open_BD_worksheet(self.open_workbook(name)))
        else:
            print("primeiro carregue uma planilha")
    def popupbutton_open_excel(self, popup, listbox): 
        book_name = self.get_selected_option(popup, listbox)
        self.this_book_name = book_name
        self.insert_tree(self.open_BD_worksheet(self.open_workbook(book_name)))
    def monthly_report(self) -> None:
        try:
            name = self.this_book_name
        except AttributeError:
            return messagebox.showerror("Erro", "Por favor abra uma planilha primeiro")
        quest = messagebox.askyesno("Relatorio", "Deseja realmente fazer o relatorio ?")
        if quest is True:
            wb = self.open_workbook(name)
            bd_ws = self.open_BD_worksheet(wb)
            m_ws = self.open_monthly_worksheet(wb)
            dados = self.make_monthly_data(bd_ws)
            self.insert_monthly_data(m_ws, dados, name)
            wb.save(name)

class Window(Buttons):
    def __init__(self):
        self.create_window()
        self.create_frames()
        self.label_frame()
        self.entry_label()
        self.output_list()
        self.buttons()
        self.menu()
    def create_window(self):
        root.title("Impressão Digital")
        root.geometry("900x700")
        root.resizable(False, False)
        root.configure(background= "#700316")
        root.grid_rowconfigure(0, weight=1)
        root.grid_columnconfigure(0, weight=1)
    def create_frames(self):
        self.frame_1 = Frame(root, bg="white", width="880", height="350", highlightbackground="#70032c", highlightthickness=6)
        self.frame_1.grid(column=0, row=0, sticky="n")
        self.frame_2 = Frame(root, bg="white", width="880", height="340", highlightbackground="#70032c", highlightthickness=6)
        self.frame_2.grid(column=0, row=1, sticky="s")
    def label_frame(self):
        self.lf_filter = LabelFrame(self.frame_1, text="Pesquisar", borderwidth=1, relief="solid")
        self.lf_filter.place(x=15, y=40, width=390, height=50)
        self.lf_dados = LabelFrame(self.frame_1, text="Inserir Dados", borderwidth=1, relief="solid")
        self.lf_dados.place(x=15, y=100, width=500, height=220)
        self.lf_pagamentos = LabelFrame(self.lf_dados, text="Pagamentos", borderwidth=1, relief="solid")
        self.lf_pagamentos.place(x=300, y=40, width=185, height=120)
        self.lf_pagamentos = LabelFrame(self.lf_dados, text="Impressão", borderwidth=1, relief="solid")
        self.lf_pagamentos.place(x=15, y=40, width=250, height=120)     
    def buttons(self):

        self.bt_buscar = Button(self.frame_1, text="Buscar", command=self.button_find)
        self.bt_buscar.place(relx=0.24, rely=0.15, relheight=0.1, relwidth=0.1)

        self.bt_apagar = Button(self.frame_1, text="Apagar", )
        self.bt_apagar.place(relx=0.35, rely=0.15, relheight=0.1, relwidth=0.1)

        self.bt_salvar = Button(self.frame_1, text="Salvar", command=self.button_save)
        self.bt_salvar.place(relx=0.036, rely=0.83, relheight=0.1, relwidth=0.1)

        self.bt_limpar = Button(self.frame_1, text="Limpar", command=self.clean_table)
        self.bt_limpar.place(relx=0.48, rely=0.83, relheight=0.1, relwidth=0.1)
    def entry_label(self):

        self.label_id = Label(self.frame_1, text="ID")
        self.label_id.place(relx=0.03, rely=0.18, relheight=0.05, relwidth=0.05)
        self.entry_id = Entry(self.frame_1)
        self.entry_id.place(relx=0.11, rely=0.15, relheight=0.1, relwidth=0.08)

        self.label_data = Label(self.frame_1, text="Data")
        self.label_data.place(relx=0.04, rely=0.38, relheight=0.05, relwidth=0.05)
        self.entry_data = DateEntry(self.frame_1, selectmode="day")
        self.entry_data.place(relx=0.11, rely=0.35, relheight=0.1, relwidth=0.08)

        self.label_hora = Label(self.frame_1, text="Hora")
        self.label_hora.place(relx=0.22, rely=0.38, relheight=0.05, relwidth=0.05)
        self.entry_hora = Entry(self.frame_1)
        self.entry_hora.place(relx=0.3, rely=0.35, relheight=0.1, relwidth=0.08)

        self.label_copias = Label(self.frame_1, text="Cópias")
        self.label_copias.place(relx=0.05, rely=0.58, relheight=0.05, relwidth=0.05)
        self.label_perdas = Label(self.frame_1, text="Perdas")
        self.label_perdas.place(relx=0.05, rely=0.68, relheight=0.05, relwidth=0.05) 
        self.label_copias_br = Label(self.frame_1, text="Brother")
        self.label_copias_br.place(relx=0.12, rely=0.5, relheight=0.05, relwidth=0.05)
        self.label_copias_r = Label(self.frame_1, text="Ricoh")
        self.label_copias_r.place(relx=0.215, rely=0.5, relheight=0.05, relwidth=0.05) 
        self.entry_copias_br = Entry(self.frame_1)
        self.entry_copias_br.place(relx=0.12, rely=0.56, relheight=0.1, relwidth=0.08)
        self.entry_copias_r = Entry(self.frame_1)
        self.entry_copias_r.place(relx=0.22, rely=0.56, relheight=0.1, relwidth=0.08)
        self.entry_perdas_br = Entry(self.frame_1)
        self.entry_perdas_br.place(relx=0.12, rely=0.66, relheight=0.1, relwidth=0.08)
        self.entry_perdas_r = Entry(self.frame_1)
        self.entry_perdas_r.place(relx=0.22, rely=0.66, relheight=0.1, relwidth=0.08)

        self.label_pg_dinheiro = Label(self.frame_1, text="Dinheiro")
        self.label_pg_dinheiro.place(relx=0.38, rely=0.58, relheight=0.05, relwidth=0.06)
        self.label_pg_pix = Label(self.frame_1, text="Pix")
        self.label_pg_pix.place(relx=0.38, rely=0.68, relheight=0.05, relwidth=0.06)
        self.entry_pg_dinheiro = Entry(self.frame_1)
        self.entry_pg_dinheiro.place(relx=0.46, rely=0.56, relheight=0.1, relwidth=0.08)
        self.entry_pg_pix = Entry(self.frame_1)
        self.entry_pg_pix.place(relx=0.46, rely=0.66, relheight=0.1, relwidth=0.08)
    def output_list(self):
        self.list_print = ttk.Treeview(self.frame_2, height=5, columns=("col1","col2","col3","col4","col5","col6","col7","col8","col9"))
        self.list_print.place(relx=0.01, rely=0.1, relwidth=0.95, relheight=0.85)
        
        self.list_scroll = Scrollbar(self.frame_2, orient="vertical")
        self.list_print.configure(yscroll=self.list_scroll.set)
        self.list_scroll.place(relx=0.96,rely=0.1,relheight=0.85,relwidth=0.035)

        self.list_print.heading("#0", text="")
        self.list_print.heading("#1", text="ID")
        self.list_print.heading("#2", text="Data")
        self.list_print.heading("#3", text="Hora")
        self.list_print.heading("#4", text="Brother")
        self.list_print.heading("#5", text="Ricoh")
        self.list_print.heading("#6", text="Brother")
        self.list_print.heading("#7", text="Ricoh")
        self.list_print.heading("#8", text="Dinheiro")
        self.list_print.heading("#9", text="Pix")

        self.list_print.column("#0", width=1)
        self.list_print.column("#1", width=50)
        self.list_print.column("#2", width=80)
        self.list_print.column("#3", width=50)
        self.list_print.column("#4", width=60)
        self.list_print.column("#5", width=60)
        self.list_print.column("#6", width=60)
        self.list_print.column("#7", width=60)
        self.list_print.column("#8", width=60)
        self.list_print.column("#9", width=60)

        self.list_print.bind('<Double-Button-1>', self.on_double_click)

        self.frame2_label_perdas = Label(self.frame_2, text="Copias", borderwidth=2, relief="solid")
        self.frame2_label_perdas.place(relx=0.344, rely=0.001, relheight=0.1 ,relwidth=0.206)
        self.frame2_label_totais = Label(self.frame_2, text="Perdas", borderwidth=2, relief="solid")
        self.frame2_label_totais.place(relx=0.550, rely=0.001, relheight=0.1 ,relwidth=0.201)
        self.frame2_label_dinheiro = Label(self.frame_2, text="Pagamento", borderwidth=2, relief="solid")
        self.frame2_label_dinheiro.place(relx=0.751, rely=0.001, relheight=0.1 ,relwidth=0.208)
    def menu(self):
        menubar = Menu(root)
        root.config(menu=menubar)
        filemenu = Menu(menubar)
        filemenu2 = Menu(menubar)
        def Quit(): root.destroy()
        menubar.add_cascade(label="Opções", menu=filemenu)
        menubar.add_cascade(label="Sobre", menu=filemenu2)
        filemenu.add_command(label="Abrir Planilha", command=self.open_excel_popup)
        filemenu.add_command(label="Fazer Relatorio", command=self.monthly_report)
        filemenu.add_command(label="Sair", command=Quit)
        filemenu2.add_command(label="Limpar Cliente", command=self.clean_table)
    def open_excel_popup(self):
        files = self.find_excel_files()
        popup = Toplevel(root)
        popup.title("Carregar planilha de excel")
        label = Label(popup, text="Selecione o mês que deseja carregar:")
        label.pack()
        listbox = Listbox(popup)
        listbox.pack()
        for file in files:
            listbox.insert(END, file)
        listbox.selection_set(0)
        button = Button(popup, text="OK", command=lambda: self.popupbutton_open_excel(popup, listbox))
        button.pack()  
        popup.wait_window()

def main():
    Window()
    Aplication()
    root.mainloop()

if __name__ == "__main__":
    main()